#!/usr/bin/env python3
"""
Расширенный генератор отчетов с поддержкой:
- Сворачиваемых колонок и строк
- Автофильтров
- Условного форматирования
- Многоуровневой группировки
- Интерактивных элементов
- Изображений в таблицах и секциях
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import json
from jinja2 import Template
import tempfile
import os
import requests
import base64
from io import BytesIO
from PIL import Image as PILImage, ImageDraw, ImageFont


class AdvancedExcelRenderer:
    """Расширенный рендерер Excel с продвинутыми возможностями"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        self.styles_created = False
        
    def create_styles(self):
        """Создание именованных стилей"""
        if self.styles_created:
            return
            
        # Стиль заголовка
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, size=14, color="FFFFFF")
        header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Стиль подзаголовка
        subheader_style = NamedStyle(name="subheader_style")
        subheader_style.font = Font(bold=True, size=12, color="000000")
        subheader_style.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        subheader_style.alignment = Alignment(horizontal="left", vertical="center")
        
        # Стиль данных
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(size=10)
        data_style.alignment = Alignment(horizontal="left", vertical="center")
        data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Стиль числовых данных
        number_style = NamedStyle(name="number_style")
        number_style.font = Font(size=10)
        number_style.alignment = Alignment(horizontal="right", vertical="center")
        number_style.number_format = '#,##0.00'
        number_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Добавляем стили в книгу
        try:
            self.wb.add_named_style(header_style)
            self.wb.add_named_style(subheader_style)
            self.wb.add_named_style(data_style)
            self.wb.add_named_style(number_style)
        except ValueError:
            # Стили уже существуют
            pass
            
        self.styles_created = True
    
    def create_collapsible_report(self, data, template_config=None):
        """
        Создание отчета со сворачиваемыми секциями
        
        Args:
            data: Данные для отчета
            template_config: Конфигурация шаблона
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Сложный отчет"
        
        self.create_styles()
        
        current_row = 1
        
        # Заголовок отчета
        current_row = self._add_report_header(data, current_row)
        
        # Основные метрики (всегда видимые)
        if 'summary' in data:
            current_row = self._add_summary_section(data['summary'], current_row)
        
        # Детальные секции (сворачиваемые)
        if 'sections' in data:
            for section in data['sections']:
                current_row = self._add_collapsible_section(section, current_row)
        
        # Добавляем автофильтры и форматирование
        self._apply_advanced_formatting()
        
        return self.wb
    
    def _add_report_header(self, data, start_row):
        """Добавление заголовка отчета"""
        title = data.get('title', 'Сложный отчет')
        subtitle = data.get('subtitle', f"Создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        
        # Основной заголовок
        self.ws.cell(row=start_row, column=1, value=title)
        self.ws.cell(row=start_row, column=1).style = "header_style"
        self.ws.merge_cells(f'A{start_row}:F{start_row}')
        
        # Подзаголовок
        self.ws.cell(row=start_row + 1, column=1, value=subtitle)
        self.ws.cell(row=start_row + 1, column=1).style = "subheader_style"
        self.ws.merge_cells(f'A{start_row + 1}:F{start_row + 1}')
        
        return start_row + 3
    
    def _add_summary_section(self, summary_data, start_row):
        """Добавление секции сводки"""
        # Заголовок секции
        self.ws.cell(row=start_row, column=1, value="📊 ОСНОВНЫЕ ПОКАЗАТЕЛИ")
        self.ws.cell(row=start_row, column=1).style = "subheader_style"
        self.ws.merge_cells(f'A{start_row}:F{start_row}')
        
        current_row = start_row + 1
        
        # Метрики в виде карточек
        col = 1
        for key, value in summary_data.items():
            if col > 6:  # Переход на новую строку
                current_row += 1
                col = 1
            
            # Название метрики
            self.ws.cell(row=current_row, column=col, value=key.replace('_', ' ').title())
            self.ws.cell(row=current_row, column=col).font = Font(bold=True, size=9)
            
            # Значение метрики
            self.ws.cell(row=current_row + 1, column=col, value=value)
            self.ws.cell(row=current_row + 1, column=col).style = "number_style"
            
            col += 1
        
        return current_row + 3
    
    def _add_collapsible_section(self, section_data, start_row):
        """Добавление сворачиваемой секции"""
        section_title = section_data.get('title', 'Секция')
        section_type = section_data.get('type', 'table')
        is_collapsed = section_data.get('collapsed', False)
        
        # Заголовок секции с кнопкой сворачивания
        collapse_symbol = "▼" if not is_collapsed else "▶"
        header_text = f"{collapse_symbol} {section_title}"
        
        self.ws.cell(row=start_row, column=1, value=header_text)
        self.ws.cell(row=start_row, column=1).style = "subheader_style"
        self.ws.merge_cells(f'A{start_row}:F{start_row}')
        
        section_start_row = start_row + 1
        
        if section_type == 'table':
            current_row = self._add_table_with_filters(section_data, section_start_row)
        elif section_type == 'grouped_data':
            current_row = self._add_grouped_data(section_data, section_start_row)
        elif section_type == 'chart':
            current_row = self._add_chart_section(section_data, section_start_row)
        elif section_type == 'image':
            current_row = self._add_image_section(section_data, section_start_row)
        elif section_type == 'drawing':
            current_row = self._add_drawing_section(section_data, section_start_row)
        else:
            current_row = section_start_row
        
        # Настройка группировки для сворачивания
        if current_row > section_start_row:
            self._setup_row_grouping(section_start_row, current_row - 1, hidden=is_collapsed)
        
        return current_row + 1
    
    def _add_table_with_filters(self, section_data, start_row):
        """Добавление таблицы с автофильтрами"""
        data = section_data.get('data', [])
        image_columns = section_data.get('image_columns', [])
        
        if not data:
            return start_row
        
        # Проверяем, что data - это список словарей
        if not isinstance(data, list):
            print(f"Предупреждение: неверный формат данных в секции '{section_data.get('title', 'Unknown')}' - ожидается список")
            return start_row
            
        if not data or not isinstance(data[0], dict):
            print(f"Предупреждение: пустые данные или неверный формат в секции '{section_data.get('title', 'Unknown')}'")
            return start_row

        # Если есть колонки с изображениями, используем специальный метод
        if image_columns:
            return self._add_table_with_images(section_data, start_row)

        df = pd.DataFrame(data)
        
        # Заголовки таблицы
        for col_idx, column in enumerate(df.columns, 1):
            cell = self.ws.cell(row=start_row, column=col_idx, value=column)
            cell.style = "header_style"
        
        # Данные таблицы
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.style = "number_style"
                else:
                    cell.style = "data_style"
        
        # Создание таблицы Excel с автофильтрами
        table_range = f"A{start_row}:{chr(64 + len(df.columns))}{start_row + len(df)}"
        table = Table(displayName=f"Table{start_row}", ref=table_range)
        
        # Стиль таблицы
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        self.ws.add_table(table)
        
        # Условное форматирование для числовых колонок
        self._apply_conditional_formatting(df, start_row)
        
        return start_row + len(df) + 1
    
    def _add_grouped_data(self, section_data, start_row):
        """Добавление группированных данных с многоуровневым сворачиванием"""
        groups = section_data.get('groups', [])
        
        if not isinstance(groups, list):
            print(f"Предупреждение: неверный формат групп в секции '{section_data.get('title', 'Unknown')}' - ожидается список")
            return start_row
        
        current_row = start_row
        
        for group in groups:
            if not isinstance(group, dict):
                print(f"Предупреждение: неверный формат группы в секции '{section_data.get('title', 'Unknown')}'")
                continue
                
            group_title = group.get('title', 'Группа')
            group_data = group.get('data', [])
            is_collapsed = group.get('collapsed', False)
            
            # Заголовок группы (уровень 1)
            collapse_symbol = "▼" if not is_collapsed else "▶"
            self.ws.cell(row=current_row, column=1, value=f"  {collapse_symbol} {group_title}")
            self.ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            
            group_start_row = current_row + 1
            
            # Проверяем данные группы
            if not isinstance(group_data, list) or not group_data:
                current_row = group_start_row
                continue
                
            if not isinstance(group_data[0], dict):
                print(f"Предупреждение: неверный формат данных группы '{group_title}'")
                current_row = group_start_row
                continue
            
            # Данные группы
            df = pd.DataFrame(group_data)
            
            # Заголовки
            for col_idx, column in enumerate(df.columns, 2):  # Смещение для отступа
                cell = self.ws.cell(row=group_start_row, column=col_idx, value=column)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Данные
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), group_start_row + 1):
                for col_idx, value in enumerate(row_data, 2):
                    cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                    if isinstance(value, (int, float)):
                        cell.style = "number_style"
                    else:
                        cell.style = "data_style"
            
            current_row = group_start_row + len(df) + 1
            
            # Группировка строк группы (уровень 2)
            if current_row > group_start_row:
                self._setup_row_grouping(group_start_row, current_row - 1, level=2, hidden=is_collapsed)
            
            current_row += 1
        
        return current_row
    
    def _add_chart_section(self, section_data, start_row):
        """Добавление секции с графиком"""
        chart_type = section_data.get('chart_type', 'bar')
        data = section_data.get('data', [])
        
        # Проверяем данные
        if not isinstance(data, list) or not data:
            print(f"Предупреждение: пустые данные или неверный формат в графике '{section_data.get('title', 'Unknown')}'")
            return start_row
            
        if not isinstance(data[0], dict):
            print(f"Предупреждение: неверный формат данных графика в секции '{section_data.get('title', 'Unknown')}'")
            return start_row
        
        df = pd.DataFrame(data)
        
        # Добавляем данные для графика
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for col_idx, value in enumerate(row_data, 1):
                self.ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Создаем график
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()
        
        # Настройка данных графика
        data_range = Reference(self.ws, min_col=2, min_row=start_row + 1, 
                              max_col=len(df.columns), max_row=start_row + len(df))
        categories = Reference(self.ws, min_col=1, min_row=start_row + 1, 
                              max_row=start_row + len(df))
        
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        
        # Размещение графика
        chart.width = 15
        chart.height = 10
        self.ws.add_chart(chart, f"H{start_row}")
        
        return start_row + len(df) + 15  # Учитываем высоту графика
    
    def _setup_row_grouping(self, start_row, end_row, level=1, hidden=False):
        """Настройка группировки строк с кнопками сворачивания"""
        if start_row >= end_row:
            return
            
        # Устанавливаем уровень группировки для каждой строки
        for row in range(start_row, end_row + 1):
            self.ws.row_dimensions[row].outline_level = level
            if hidden:
                self.ws.row_dimensions[row].hidden = True
    
    def _apply_conditional_formatting(self, df, start_row):
        """Применение условного форматирования"""
        for col_idx, column in enumerate(df.columns, 1):
            if df[column].dtype in ['int64', 'float64']:
                # Цветовая шкала для числовых данных
                range_str = f"{chr(64 + col_idx)}{start_row + 1}:{chr(64 + col_idx)}{start_row + len(df)}"
                rule = ColorScaleRule(
                    start_type='min', start_color='F8696B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                    end_type='max', end_color='63BE7B'
                )
                self.ws.conditional_formatting.add(range_str, rule)
    
    def _apply_advanced_formatting(self):
        """Применение продвинутого форматирования"""
        # Автоподбор ширины колонок
        for column in self.ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                # Пропускаем объединенные ячейки
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                self.ws.column_dimensions[column_letter].width = adjusted_width
        
        # Закрепление области
        self.ws.freeze_panes = 'A4'
    
    def _add_image_section(self, section_data, start_row):
        """Добавление секции с изображением"""
        image_config = section_data.get('image_config', {})
        
        if not image_config:
            print(f"Предупреждение: отсутствует конфигурация изображения в секции '{section_data.get('title', 'Unknown')}'")
            return start_row + 2
        
        image_type = image_config.get('type', 'url')
        source = image_config.get('source', '')
        width = image_config.get('width', 300)
        height = image_config.get('height', 200)
        anchor = image_config.get('anchor', f'B{start_row}')
        description = image_config.get('description', '')
        
        try:
            excel_img = None
            
            if image_type == 'url':
                excel_img = self._load_image_from_url(source)
            elif image_type == 'base64':
                excel_img = self._load_image_from_base64(source)
            elif image_type == 'file':
                excel_img = self._load_image_from_file(source)
            
            if excel_img:
                # Устанавливаем размеры
                excel_img.width = width
                excel_img.height = height
                
                # Добавляем изображение в лист
                self.ws.add_image(excel_img, anchor)
                
                # Добавляем описание если есть
                if description:
                    desc_row = start_row + int(height / 20) + 1
                    self.ws.cell(row=desc_row, column=2, value=description)
                    self.ws.cell(row=desc_row, column=2).font = Font(italic=True, size=9)
                
                return start_row + int(height / 20) + 3
            else:
                # Если изображение не загрузилось, добавляем текст-заглушку
                self.ws.cell(row=start_row, column=2, value=f"❌ Не удалось загрузить изображение: {source}")
                self.ws.cell(row=start_row, column=2).font = Font(color="FF0000")
                return start_row + 2
                
        except Exception as e:
            print(f"Ошибка при добавлении изображения: {e}")
            self.ws.cell(row=start_row, column=2, value=f"❌ Ошибка загрузки изображения: {str(e)}")
            self.ws.cell(row=start_row, column=2).font = Font(color="FF0000")
            return start_row + 2
    
    def _add_drawing_section(self, section_data, start_row):
        """Добавление секции с программным рисованием"""
        drawing_config = section_data.get('drawing_config', {})
        drawing_type = drawing_config.get('type', 'diagram')
        
        if drawing_type == 'diagram':
            return self._create_diagram(drawing_config, start_row)
        elif drawing_type == 'flowchart':
            return self._create_flowchart(drawing_config, start_row)
        elif drawing_type == 'infographic':
            return self._create_infographic(drawing_config, start_row)
        elif drawing_type == 'custom':
            return self._create_custom_drawing(drawing_config, start_row)
        else:
            print(f"Неподдерживаемый тип рисования: {drawing_type}")
            return start_row + 2
    
    def _load_image_from_url(self, url):
        """Загрузка изображения из URL"""
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            return Image(BytesIO(response.content))
        except Exception as e:
            print(f"Ошибка загрузки изображения из URL {url}: {e}")
            return None
    
    def _load_image_from_base64(self, base64_string):
        """Загрузка изображения из base64 строки"""
        try:
            # Убираем префикс data:image/...;base64, если есть
            if ',' in base64_string:
                base64_string = base64_string.split(',')[1]
            
            image_data = base64.b64decode(base64_string)
            return Image(BytesIO(image_data))
        except Exception as e:
            print(f"Ошибка декодирования base64 изображения: {e}")
            return None
    
    def _load_image_from_file(self, file_path):
        """Загрузка изображения из файла"""
        try:
            if os.path.exists(file_path):
                return Image(file_path)
            else:
                print(f"Файл изображения не найден: {file_path}")
                return None
        except Exception as e:
            print(f"Ошибка загрузки изображения из файла {file_path}: {e}")
            return None
    
    def _add_table_with_images(self, section_data, start_row):
        """Добавление таблицы с поддержкой изображений в ячейках"""
        data = section_data.get('data', [])
        image_columns = section_data.get('image_columns', [])  # Колонки с изображениями
        
        if not data:
            return start_row
        
        if not isinstance(data, list) or not isinstance(data[0], dict):
            print(f"Предупреждение: неверный формат данных в секции '{section_data.get('title', 'Unknown')}'")
            return start_row

        df = pd.DataFrame(data)
        
        # Заголовки таблицы
        for col_idx, column in enumerate(df.columns, 1):
            cell = self.ws.cell(row=start_row, column=col_idx, value=column)
            cell.style = "header_style"
        
        # Данные таблицы
        current_row = start_row + 1
        for row_idx, (_, row_data) in enumerate(df.iterrows()):
            row_height = 20  # Стандартная высота строки
            
            for col_idx, (column, value) in enumerate(row_data.items(), 1):
                cell = self.ws.cell(row=current_row, column=col_idx)
                
                # Проверяем, является ли колонка колонкой с изображениями
                if column in image_columns and value:
                    try:
                        # Пытаемся загрузить изображение
                        excel_img = None
                        
                        if isinstance(value, dict):
                            # Если значение - это конфигурация изображения
                            img_type = value.get('type', 'url')
                            img_source = value.get('source', '')
                            img_width = value.get('width', 80)
                            img_height = value.get('height', 60)
                            
                            if img_type == 'url':
                                excel_img = self._load_image_from_url(img_source)
                            elif img_type == 'base64':
                                excel_img = self._load_image_from_base64(img_source)
                            elif img_type == 'file':
                                excel_img = self._load_image_from_file(img_source)
                        else:
                            # Если значение - это просто URL или путь к файлу
                            if str(value).startswith('http'):
                                excel_img = self._load_image_from_url(str(value))
                            else:
                                excel_img = self._load_image_from_file(str(value))
                            
                            img_width = 80
                            img_height = 60
                        
                        if excel_img:
                            excel_img.width = img_width
                            excel_img.height = img_height
                            
                            # Добавляем изображение в ячейку
                            cell_address = f"{chr(64 + col_idx)}{current_row}"
                            self.ws.add_image(excel_img, cell_address)
                            
                            # Устанавливаем высоту строки
                            row_height = max(row_height, img_height + 10)
                            
                            # Устанавливаем ширину колонки
                            col_letter = chr(64 + col_idx)
                            current_width = self.ws.column_dimensions[col_letter].width or 10
                            new_width = max(current_width, (img_width / 7) + 2)
                            self.ws.column_dimensions[col_letter].width = new_width
                        else:
                            cell.value = "❌ Изображение не загружено"
                            cell.font = Font(color="FF0000", size=8)
                            
                    except Exception as e:
                        print(f"Ошибка при добавлении изображения в ячейку: {e}")
                        cell.value = "❌ Ошибка"
                        cell.font = Font(color="FF0000", size=8)
                else:
                    # Обычные данные
                    cell.value = value
                    if isinstance(value, (int, float)):
                        cell.style = "number_style"
                    else:
                        cell.style = "data_style"
            
            # Устанавливаем высоту строки
            self.ws.row_dimensions[current_row].height = row_height
            current_row += 1
        
        return current_row + 1
    
    def _create_diagram(self, config, start_row):
        """Создание диаграммы программно"""
        diagram_data = config.get('data', [])
        diagram_style = config.get('style', 'boxes')
        
        # Создаем изображение с помощью PIL
        img_width = config.get('width', 600)
        img_height = config.get('height', 400)
        
        # Создаем изображение
        pil_img = PILImage.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(pil_img)
        
        try:
            # Пытаемся загрузить шрифт
            font = ImageFont.truetype("arial.ttf", 12)
            title_font = ImageFont.truetype("arial.ttf", 16)
        except:
            # Используем стандартный шрифт если arial не найден
            font = ImageFont.load_default()
            title_font = ImageFont.load_default()
        
        # Заголовок диаграммы
        title = config.get('title', 'Диаграмма')
        draw.text((img_width//2 - len(title)*4, 10), title, fill='black', font=title_font)
        
        if diagram_style == 'boxes':
            self._draw_box_diagram(draw, diagram_data, img_width, img_height, font)
        elif diagram_style == 'circles':
            self._draw_circle_diagram(draw, diagram_data, img_width, img_height, font)
        elif diagram_style == 'flow':
            self._draw_flow_diagram(draw, diagram_data, img_width, img_height, font)
        
        try:
            # Сохраняем изображение в память
            img_buffer = BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Добавляем в Excel
            excel_img = Image(img_buffer)
            excel_img.width = img_width // 2  # Масштабируем для Excel
            excel_img.height = img_height // 2
            
            self.ws.add_image(excel_img, f'B{start_row}')
            
        except Exception as e:
            print(f"Ошибка добавления диаграммы в Excel: {e}")
        
        return start_row + int(img_height / 30) + 2
    
    def _draw_box_diagram(self, draw, data, width, height, font):
        """Рисование диаграммы с прямоугольниками"""
        if not data:
            return
        
        box_width = min(120, (width - 100) // len(data))
        box_height = 60
        start_y = height // 2 - box_height // 2
        
        for i, item in enumerate(data):
            x = 50 + i * (box_width + 20)
            y = start_y
            
            # Цвет прямоугольника
            color = item.get('color', '#4472C4')
            if color.startswith('#'):
                color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
            else:
                color = (68, 114, 196)  # Синий по умолчанию
            
            # Рисуем прямоугольник
            draw.rectangle([x, y, x + box_width, y + box_height], 
                          fill=color, outline='black', width=2)
            
            # Текст
            text = item.get('label', f'Item {i+1}')
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]
            
            text_x = x + (box_width - text_width) // 2
            text_y = y + (box_height - text_height) // 2
            
            draw.text((text_x, text_y), text, fill='white', font=font)
            
            # Значение под прямоугольником
            value = item.get('value', '')
            if value:
                value_text = str(value)
                value_bbox = draw.textbbox((0, 0), value_text, font=font)
                value_width = value_bbox[2] - value_bbox[0]
                value_x = x + (box_width - value_width) // 2
                draw.text((value_x, y + box_height + 5), value_text, fill='black', font=font)
    
    def _draw_circle_diagram(self, draw, data, width, height, font):
        """Рисование диаграммы с кругами"""
        if not data:
            return
        
        circle_radius = min(40, min(width, height) // 6)
        center_y = height // 2
        
        for i, item in enumerate(data):
            x = 80 + i * (circle_radius * 3)
            y = center_y
            
            # Цвет круга
            color = item.get('color', '#4472C4')
            if color.startswith('#'):
                color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
            else:
                color = (68, 114, 196)
            
            # Рисуем круг
            draw.ellipse([x - circle_radius, y - circle_radius, 
                         x + circle_radius, y + circle_radius], 
                        fill=color, outline='black', width=2)
            
            # Текст в центре круга
            text = item.get('label', f'Item {i+1}')
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]
            
            text_x = x - text_width // 2
            text_y = y - text_height // 2
            
            draw.text((text_x, text_y), text, fill='white', font=font)
            
            # Значение под кругом
            value = item.get('value', '')
            if value:
                value_text = str(value)
                value_bbox = draw.textbbox((0, 0), value_text, font=font)
                value_width = value_bbox[2] - value_bbox[0]
                value_x = x - value_width // 2
                draw.text((value_x, y + circle_radius + 10), value_text, fill='black', font=font)
    
    def _draw_flow_diagram(self, draw, data, width, height, font):
        """Рисование блок-схемы"""
        if not data:
            return
        
        box_width = min(150, (width - 100) // len(data))
        box_height = 50
        start_y = height // 2 - box_height // 2
        
        for i, item in enumerate(data):
            x = 50 + i * (box_width + 30)
            y = start_y
            
            # Цвет блока
            color = item.get('color', '#4472C4')
            if color.startswith('#'):
                color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
            else:
                color = (68, 114, 196)
            
            # Рисуем блок
            draw.rectangle([x, y, x + box_width, y + box_height], 
                          fill=color, outline='black', width=2)
            
            # Текст
            text = item.get('label', f'Step {i+1}')
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]
            
            text_x = x + (box_width - text_width) // 2
            text_y = y + (box_height - text_height) // 2
            
            draw.text((text_x, text_y), text, fill='white', font=font)
            
            # Стрелка к следующему блоку
            if i < len(data) - 1:
                arrow_start_x = x + box_width
                arrow_end_x = x + box_width + 30
                arrow_y = y + box_height // 2
                
                # Линия стрелки
                draw.line([arrow_start_x, arrow_y, arrow_end_x, arrow_y], 
                         fill='black', width=2)
                
                # Наконечник стрелки
                draw.polygon([arrow_end_x, arrow_y, 
                             arrow_end_x - 10, arrow_y - 5,
                             arrow_end_x - 10, arrow_y + 5], 
                            fill='black')
    
    def _create_flowchart(self, config, start_row):
        """Создание блок-схемы"""
        # Пока используем тот же метод что и для диаграмм
        return self._create_diagram(config, start_row)
    
    def _create_infographic(self, config, start_row):
        """Создание инфографики"""
        infographic_data = config.get('data', [])
        
        # Создаем изображение
        img_width = config.get('width', 800)
        img_height = config.get('height', 500)
        
        pil_img = PILImage.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(pil_img)
        
        try:
            font = ImageFont.truetype("arial.ttf", 12)
            title_font = ImageFont.truetype("arial.ttf", 16)
        except:
            font = ImageFont.load_default()
            title_font = ImageFont.load_default()
        
        # Заголовок
        title = config.get('title', 'Инфографика')
        draw.text((img_width//2 - len(title)*6, 10), title, fill='black', font=title_font)
        
        # Размещение элементов
        cols = 3  # Количество колонок
        rows = (len(infographic_data) + cols - 1) // cols
        
        cell_width = img_width // cols
        cell_height = (img_height - 50) // rows
        
        for i, item in enumerate(infographic_data):
            col = i % cols
            row = i // cols
            
            x = col * cell_width + 20
            y = 50 + row * cell_height + 20
            
            item_type = item.get('type', 'metric')
            
            if item_type == 'metric':
                self._draw_metric_card(draw, item, x, y, font, title_font)
            elif item_type == 'progress':
                self._draw_progress_bar(draw, item, x, y, font)
            elif item_type == 'icon':
                self._draw_icon_with_text(draw, item, x, y, font)
        
        try:
            # Сохраняем изображение в память
            img_buffer = BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Добавляем в Excel
            excel_img = Image(img_buffer)
            excel_img.width = img_width // 2
            excel_img.height = img_height // 2
            
            self.ws.add_image(excel_img, f'B{start_row}')
            
        except Exception as e:
            print(f"Ошибка добавления инфографики в Excel: {e}")
        
        return start_row + int(img_height / 30) + 2
    
    def _draw_metric_card(self, draw, item, x, y, font, title_font):
        """Рисование карточки метрики"""
        card_width = 200
        card_height = 80
        
        # Цвет карточки
        color = item.get('color', '#4472C4')
        if color.startswith('#'):
            color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
        else:
            color = (68, 114, 196)
        
        # Рисуем карточку
        draw.rectangle([x, y, x + card_width, y + card_height], 
                      fill=color, outline='black', width=1)
        
        # Заголовок
        label = item.get('label', 'Метрика')
        draw.text((x + 10, y + 10), label, fill='white', font=font)
        
        # Значение
        value = item.get('value', '0')
        draw.text((x + 10, y + 35), str(value), fill='white', font=title_font)
    
    def _draw_progress_bar(self, draw, item, x, y, font):
        """Рисование прогресс-бара"""
        bar_width = 200
        bar_height = 20
        progress = item.get('progress', 0)
        
        # Фон прогресс-бара
        draw.rectangle([x, y + 20, x + bar_width, y + 20 + bar_height], 
                      fill='lightgray', outline='black', width=1)
        
        # Заполненная часть
        filled_width = int(bar_width * progress / 100)
        color = item.get('color', '#70AD47')
        if color.startswith('#'):
            color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
        else:
            color = (112, 173, 71)
        
        if filled_width > 0:
            draw.rectangle([x, y + 20, x + filled_width, y + 20 + bar_height], 
                          fill=color, outline='black', width=1)
        
        # Подпись
        label = item.get('label', 'Прогресс')
        draw.text((x, y), label, fill='black', font=font)
        
        # Процент
        draw.text((x + bar_width + 10, y + 20), f"{progress}%", fill='black', font=font)
    
    def _draw_icon_with_text(self, draw, item, x, y, font):
        """Рисование иконки с текстом"""
        symbol = item.get('symbol', '●')
        text = item.get('text', 'Текст')
        color = item.get('color', '#FFC000')
        
        if color.startswith('#'):
            color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
        else:
            color = (255, 192, 0)
        
        # Рисуем символ
        try:
            symbol_font = ImageFont.truetype("arial.ttf", 24)
        except:
            symbol_font = ImageFont.load_default()
        
        draw.text((x, y), symbol, fill=color, font=symbol_font)
        
        # Текст рядом с символом
        draw.text((x + 30, y + 5), text, fill='black', font=font)
    
    def _create_custom_drawing(self, config, start_row):
        """Создание пользовательского рисунка"""
        commands = config.get('commands', [])
        
        # Создаем изображение
        img_width = config.get('width', 500)
        img_height = config.get('height', 300)
        
        pil_img = PILImage.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(pil_img)
        
        try:
            font = ImageFont.truetype("arial.ttf", 12)
        except:
            font = ImageFont.load_default()
        
        # Выполняем команды рисования
        for cmd in commands:
            cmd_type = cmd.get('type', '')
            
            if cmd_type == 'rectangle':
                coords = cmd.get('coords', [0, 0, 100, 100])
                color = cmd.get('color', '#4472C4')
                if color.startswith('#'):
                    color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
                draw.rectangle(coords, fill=color, outline='black', width=2)
                
            elif cmd_type == 'circle':
                coords = cmd.get('coords', [0, 0, 100, 100])
                color = cmd.get('color', '#70AD47')
                if color.startswith('#'):
                    color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
                draw.ellipse(coords, fill=color, outline='black', width=2)
                
            elif cmd_type == 'line':
                coords = cmd.get('coords', [0, 0, 100, 100])
                color = cmd.get('color', 'black')
                width = cmd.get('width', 1)
                draw.line(coords, fill=color, width=width)
                
            elif cmd_type == 'text':
                position = cmd.get('position', [0, 0])
                text = cmd.get('text', 'Текст')
                color = cmd.get('color', 'black')
                size = cmd.get('size', 12)
                
                try:
                    text_font = ImageFont.truetype("arial.ttf", size)
                except:
                    text_font = font
                
                draw.text(position, text, fill=color, font=text_font)
        
        try:
            # Сохраняем изображение в память
            img_buffer = BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Добавляем в Excel
            excel_img = Image(img_buffer)
            excel_img.width = img_width // 2
            excel_img.height = img_height // 2
            
            self.ws.add_image(excel_img, f'B{start_row}')
            
        except Exception as e:
            print(f"Ошибка добавления пользовательского рисунка в Excel: {e}")
        
        return start_row + int(img_height / 30) + 2

    def save_report(self, filename):
        """Сохранение отчета"""
        self.wb.save(filename)
        return filename


def create_complex_report_template():
    """Создание шаблона для сложного отчета"""
    template_data = {
        "title": "{{report_title}}",
        "subtitle": "Период: {{period_start}} - {{period_end}} | Создан: {{creation_date}}",
        "summary": {
            "total_revenue": "{{summary.total_revenue}}",
            "total_orders": "{{summary.total_orders}}",
            "avg_order_value": "{{summary.avg_order_value}}",
            "growth_rate": "{{summary.growth_rate}}",
            "profit_margin": "{{summary.profit_margin}}",
            "customer_count": "{{summary.customer_count}}"
        },
        "sections": [
            {
                "title": "Продажи по регионам",
                "type": "table",
                "collapsed": False,
                "data": "{{regional_sales}}"
            },
            {
                "title": "Анализ продуктов",
                "type": "grouped_data",
                "collapsed": True,
                "groups": "{{product_groups}}"
            },
            {
                "title": "Динамика продаж",
                "type": "chart",
                "chart_type": "line",
                "collapsed": False,
                "data": "{{sales_dynamics}}"
            },
            {
                "title": "Детальная аналитика",
                "type": "table",
                "collapsed": True,
                "data": "{{detailed_analytics}}"
            }
        ]
    }
    
    # Сохраняем конфигурацию шаблона
    with open('complex_report_template.json', 'w', encoding='utf-8') as f:
        json.dump(template_data, f, ensure_ascii=False, indent=2)
    
    return template_data


def generate_sample_data():
    """Генерация примерных данных для демонстрации"""
    import random
    
    # Основные метрики
    summary = {
        "total_revenue": 2500000,
        "total_orders": 1250,
        "avg_order_value": 2000,
        "growth_rate": 15.5,
        "profit_margin": 28.5,
        "customer_count": 850
    }
    
    # Продажи по регионам
    regional_sales = [
        {"region": "Москва", "sales": 850000, "orders": 425, "growth": 18.2},
        {"region": "СПб", "sales": 620000, "orders": 310, "growth": 12.5},
        {"region": "Екатеринбург", "sales": 380000, "orders": 190, "growth": 8.7},
        {"region": "Новосибирск", "sales": 320000, "orders": 160, "growth": 22.1},
        {"region": "Казань", "sales": 280000, "orders": 140, "growth": 5.3},
        {"region": "Нижний Новгород", "sales": 50000, "orders": 25, "growth": -2.1}
    ]
    
    # Группированные данные по продуктам
    product_groups = [
        {
            "title": "Электроника",
            "collapsed": False,
            "data": [
                {"product": "Смартфоны", "sales": 450000, "units": 150, "margin": 25.5},
                {"product": "Ноутбуки", "sales": 380000, "units": 95, "margin": 18.2},
                {"product": "Планшеты", "sales": 220000, "units": 110, "margin": 22.1}
            ]
        },
        {
            "title": "Одежда",
            "collapsed": True,
            "data": [
                {"product": "Куртки", "sales": 180000, "units": 360, "margin": 45.2},
                {"product": "Джинсы", "sales": 150000, "units": 300, "margin": 38.5},
                {"product": "Футболки", "sales": 120000, "units": 600, "margin": 52.1}
            ]
        },
        {
            "title": "Книги",
            "collapsed": True,
            "data": [
                {"product": "Художественная литература", "sales": 85000, "units": 850, "margin": 35.2},
                {"product": "Техническая литература", "sales": 65000, "units": 325, "margin": 28.5},
                {"product": "Детские книги", "sales": 45000, "units": 450, "margin": 42.1}
            ]
        }
    ]
    
    # Динамика продаж
    sales_dynamics = []
    base_date = datetime(2024, 1, 1)
    for i in range(12):
        month_date = base_date + timedelta(days=30*i)
        sales_dynamics.append({
            "month": month_date.strftime("%B"),
            "sales": random.randint(180000, 250000),
            "orders": random.randint(90, 130),
            "customers": random.randint(60, 90)
        })
    
    # Детальная аналитика
    detailed_analytics = []
    for i in range(50):
        detailed_analytics.append({
            "order_id": f"ORD-{1000+i}",
            "customer": f"Клиент {i+1}",
            "product": random.choice(["Смартфон", "Ноутбук", "Куртка", "Книга"]),
            "amount": random.randint(500, 5000),
            "date": (datetime.now() - timedelta(days=random.randint(1, 90))).strftime("%d.%m.%Y"),
            "status": random.choice(["Выполнен", "В обработке", "Отменен"])
        })
    
    return {
        "report_title": "Комплексный отчет по продажам",
        "period_start": "01.01.2024",
        "period_end": "31.12.2024",
        "creation_date": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "summary": summary,
        "regional_sales": regional_sales,
        "product_groups": product_groups,
        "sales_dynamics": sales_dynamics,
        "detailed_analytics": detailed_analytics
    }


def render_template_with_data(template_data, context_data):
    """Рендеринг шаблона с данными используя Jinja2"""
    
    def render_recursive(obj, context):
        """Рекурсивный рендеринг объекта"""
        if isinstance(obj, str):
            if "{{" in obj or "{%" in obj:
                try:
                    template = Template(obj)
                    return template.render(**context)
                except Exception as e:
                    print(f"Ошибка рендеринга шаблона: {e}")
                    return obj
            return obj
        elif isinstance(obj, dict):
            return {key: render_recursive(value, context) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [render_recursive(item, context) for item in obj]
        else:
            return obj
    
    # Сначала обрабатываем простые подстановки
    rendered = render_recursive(template_data, context_data)
    
    # Затем обрабатываем прямые ссылки на данные
    def resolve_data_references(obj, context):
        """Разрешение прямых ссылок на данные"""
        if isinstance(obj, str):
            # Проверяем, является ли строка ссылкой на данные
            if obj.startswith("{{") and obj.endswith("}}"):
                key = obj[2:-2].strip()
                # Поддержка вложенных ключей (например, summary.total_revenue)
                try:
                    value = context
                    for part in key.split('.'):
                        value = value[part]
                    return value
                except (KeyError, TypeError):
                    print(f"Не найден ключ: {key}")
                    return obj
            return obj
        elif isinstance(obj, dict):
            return {key: resolve_data_references(value, context) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [resolve_data_references(item, context) for item in obj]
        else:
            return obj
    
    return resolve_data_references(rendered, context_data)


