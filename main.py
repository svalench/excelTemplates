import pandas as pd
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import tempfile
import os
import json
from urllib.parse import urlparse
import requests
from io import BytesIO
from datetime import datetime


class ExcelRenderer:
    def __init__(self):
        self.wb = None
        self.ws = None

    def render_template(self, template_path, output_path, context):
        """Основной метод рендеринга шаблона"""
        self.wb = load_workbook(template_path)
        self.ws = self.wb.active

        # Обработка ячеек с Jinja2
        self._process_jinja_cells(context)

        # Обработка специальных инструкций
        self._process_instructions(context)

        self.wb.save(output_path)
        return output_path

    def _process_jinja_cells(self, context):
        """Обработка ячеек с Jinja2 разметкой"""
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Обработка обычных Jinja2 выражений
                    if "{{" in cell.value or "{%" in cell.value:
                        try:
                            template = Template(cell.value)
                            rendered = template.render(**context)
                            self._set_cell_value(cell, rendered)
                        except Exception as e:
                            cell.value = f"ERROR: {str(e)}"

                    # Обработка специальных инструкций
                    elif cell.value.strip().startswith("{%excel"):
                        self._process_excel_instruction(cell, context)

    def _process_excel_instruction(self, cell, context):
        """Обработка специальных инструкций для Excel"""
        instruction = cell.value.strip()

        # Вставка изображения
        if instruction.startswith("{%excel image"):
            img_var = instruction.split('"')[1]
            img_path = context.get(img_var)
            if img_path:
                self._insert_image(cell, img_path)
            cell.value = None

        # Группировка строк
        elif instruction.startswith("{%excel group"):
            params = instruction.split('"')[1::2]
            self._group_rows(cell.row, *params)
            cell.value = None

    def _set_cell_value(self, cell, value):
        """Установка значения ячейки с сохранением типа данных"""
        if isinstance(value, str) and value.startswith("http"):
            cell.hyperlink = value
            cell.value = value
            cell.style = "Hyperlink"
        elif isinstance(value, (int, float)):
            cell.value = value
        else:
            cell.value = str(value)

    def _insert_image(self, cell, img_source):
        """Вставка изображения в указанную ячейку"""
        try:
            if img_source.startswith(('http://', 'https://')):
                response = requests.get(img_source)
                img = Image(BytesIO(response.content))
            else:
                img = Image(img_source)

            # Масштабирование изображения
            img.width = min(img.width, 300)
            img.height = min(img.height, 200)

            self.ws.add_image(img, cell.coordinate)
        except Exception as e:
            print(f"Ошибка вставки изображения: {e}")

    def _group_rows(self, start_row, end_row=None, level=1, hidden=False):
        """Группировка строк с возможностью сворачивания"""
        end_row = end_row or start_row
        for row in range(int(start_row), int(end_row) + 1):
            self.ws.row_dimensions[row].outline_level = int(level)
            if hidden:
                self.ws.row_dimensions[row].hidden = True

    def _process_instructions(self, context):
        """Обработка структурных инструкций"""
        for key, value in context.items():
            if isinstance(value, pd.DataFrame):
                self._insert_dataframe(key, value)
            elif isinstance(value, dict) and value.get('_type') == 'nested_table':
                self._insert_nested_table(value)

    def _insert_dataframe(self, marker, df):
        """Вставка DataFrame в указанное место"""
        marker_found = False
        start_row, start_col = None, None

        # Ищем маркер во всех ячейках
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == f"{{{{{marker}_start}}}}":
                    start_row, start_col = cell.row, cell.column
                    marker_found = True
                    break
            if marker_found:
                break

        if marker_found:
            # Очищаем ячейку с маркером
            self.ws.cell(row=start_row, column=start_col).value = None

            # Вставка данных
            for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row_data):
                    if value is not None:  # Пропускаем None значения
                        cell = self.ws.cell(row=start_row + r_idx, column=start_col + c_idx)
                        self._set_cell_value(cell, value)

            # Группировка если нужно
            group_marker = f"{{{{{marker}_group}}}}"
            for row in self.ws.iter_rows():
                for cell in row:
                    if cell.value == group_marker:
                        self._group_rows(start_row + 1, start_row + len(df))
                        cell.value = None
                        break

    def _insert_nested_table(self, data):
        """Вставка вложенной таблицы со сворачиванием"""
        marker = data.get('marker')
        df = pd.DataFrame(data['data'])

        # Находим родительскую строку
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == f"{{{{{marker}}}}}":
                    parent_row = cell.row

                    # Вставляем таблицу ниже
                    insert_row = parent_row + 1
                    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), insert_row):
                        for c_idx, value in enumerate(row_data, cell.column):
                            self._set_cell_value(self.ws.cell(row=r_idx, column=c_idx), value)

                    # Настраиваем группировку
                    self._group_rows(insert_row, insert_row + len(df) - 1, level=2)

                    # Добавляем "+/-" для сворачивания
                    self.ws.cell(row=parent_row, column=cell.column + 1, value="+/-")
                    return


def generate_report(template_name, json_data, output_format='xlsx'):
    """Генерация отчета с расширенными возможностями"""
    renderer = ExcelRenderer()

    # Подготовка контекста
    if isinstance(json_data, str):
        context = json.loads(json_data)
    else:
        context = json_data

    # Добавляем текущую дату если не указана
    if 'report_date' not in context:
        context['report_date'] = datetime.now().strftime('%d.%m.%Y')

    # Конвертация табличных данных
    for key, value in context.items():
        if isinstance(value, list) and all(isinstance(i, dict) for i in value):
            context[key] = pd.DataFrame(value)

    # Временный файл
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        temp_xlsx = tmp.name

    # Рендеринг
    template_path = f"{template_name}.xlsx"
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Шаблон {template_path} не найден")

    renderer.render_template(template_path, temp_xlsx, context)

    # Конвертация в PDF (если установлен LibreOffice)
    if output_format == 'pdf':
        temp_pdf = temp_xlsx.replace('.xlsx', '.pdf')
        try:
            os.system(f"libreoffice --headless --convert-to pdf {temp_xlsx} --outdir {os.path.dirname(temp_pdf)}")
            if os.path.exists(temp_pdf):
                os.unlink(temp_xlsx)
                return temp_pdf
            else:
                print("Не удалось создать PDF, возвращаем Excel файл")
                return temp_xlsx
        except Exception as e:
            print(f"Ошибка конвертации в PDF: {e}")
            return temp_xlsx

    return temp_xlsx


# Функция для быстрого создания простого отчета
def create_simple_report(title, data_dict, output_path=None):
    """Создание простого отчета без шаблона"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Заголовок
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    row = 3

    def process_value(key, value, current_row, level=0):
        """Рекурсивная обработка значений"""
        indent = "  " * level

        if isinstance(value, (list, pd.DataFrame)):
            # Заголовок таблицы
            ws.cell(row=current_row, column=1, value=f"{indent}{key}:")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            # Данные таблицы
            if isinstance(value, list):
                if value and isinstance(value[0], dict):
                    # Список словарей - создаем таблицу
                    df = pd.DataFrame(value)
                elif value and isinstance(value[0], (str, int, float)):
                    # Простой список - выводим как строки
                    for item in value:
                        ws.cell(row=current_row, column=2, value=f"• {item}")
                        current_row += 1
                    return current_row
                else:
                    # Сложный список - выводим как текст
                    ws.cell(row=current_row, column=2, value=str(value))
                    current_row += 1
                    return current_row
            else:
                df = value

            # Вставка DataFrame
            for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, cell_value in enumerate(row_data):
                    if cell_value is not None:
                        ws.cell(row=current_row + r_idx, column=c_idx + 1, value=cell_value)

            current_row += len(df) + 2

        elif isinstance(value, dict):
            # Вложенный словарь
            ws.cell(row=current_row, column=1, value=f"{indent}{key}:")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for sub_key, sub_value in value.items():
                current_row = process_value(sub_key, sub_value, current_row, level + 1)

            current_row += 1

        else:
            # Простое значение
            ws.cell(row=current_row, column=1, value=f"{indent}{key}:")
            ws.cell(row=current_row, column=2, value=str(value))
            current_row += 1

        return current_row

    # Обработка всех данных
    for key, value in data_dict.items():
        row = process_value(key, value, row)

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Сохранение
    if not output_path:
        output_path = f"simple_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb.save(output_path)
    return output_path